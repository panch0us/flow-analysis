import pandas
# импортируем openpyxl для работы с таблицами xlsx
import openpyxl
# имортируем встроенный класс Counter из модуля collection, который подсчитывает количество вхождений
# элементов в списке. (Синтаксис: a = Counter(список); a.['нужный элемент списка для подсчета'])
#from collections import Counter

# указываем дирректорию расположения файлов для дальнейшей обработки
#DIRECTORY = 'C:/Users/asus/Desktop/home/test/analiz/csv/number/'
#DIRECTORY = 'C:/Users/umvd/Desktop/Анализ_потоков/csv/номер/'
DIRECTORY = 'C:/Users/panchous/PycharmProjects/data/potok_analisys/'

# Открываем xlsx файл
wb = openpyxl.Workbook()


def return_dict_all_grz():
    """Функция принимает любого количество файлов с ГРЗ формата CSV для переработки их в нумерованный словарь,
    и возвращает полный словарь с нумерованными списками ГРЗ"""
    stop = ''
    counter = 1
    list_input = []
    dict_input = {counter: list_input}
    other_grz = []
    while stop == '':
        list_input = []
        file = input('Введити название файла: ')
        try:
            result = pandas.read_csv(DIRECTORY + file, encoding='utf_16_le')
            for el in result['ГРЗ']:
                list_input.append(el)
                dict_input[counter] = list_input
                if str(el)[-2:] != str(32):
                    other_grz.append(el)
            stop = input('ПРОДОЛЖЕНИЕ [enter]\nЗАВЕРШЕНИЕ [символ + enter]')
            counter = counter + 1
        except FileNotFoundError:
            print('Файл \'' + str(file) + '\' отсутвует в дирректории C:/Users/umvd/Desktop/Анализ_потоков/csv/номер/')
            continue
        except PermissionError:
            print('Ошибка: "PermissionError", попробуйте снова')
            continue
    return dict_input, other_grz


def iteration_one_to_many_lists(dict_input_user):
    """Функция сравнивает список ГРЗ i со списком ГРЗ  b. Совпадающие ГРЗ сохраняются в новый список по названием i + b
    После чего список i сравнивается со списокм b + 1, и т.д. до последнего списка b.
    После чего список i + 1 сравнивается со списокм b = 1 + i (список, который больше на 1 списка i).
    Результат сохраняется в xlsx файл.
    Функция получает словарь, состоящий из {номера списка: списка ГРЗ} из функции return_dict_all_grz."""

    # Создаем в открытом xlsx файле новый лист
    wb.create_sheet(title='Отдельное сравнение', index=0)
    sheet = wb['Отдельное сравнение']

    # Получаем словарь вида {№: '[список ГРЗ]} из функции return_dict_all_grz (введенные списки пользователем)
    dict_input = dict_input_user
    # Считаем количество списков в ГРЗ в полученом словаре
    count_element_from_dict = len(dict_input)
    # Между собой будут сравниваться два списка, первый - i, следующий за ним - b
    i = 0
    b = 1
    # номера ячеек для заголовка (строка + столбец)
    row_number = 1
    col_number = 1
    # номера ячеек для значений (строка + столбец)
    row = 1
    column = 1

    # Выполняем цикл по количеству списков ГРЗ в словаре
    while i < count_element_from_dict:
        i = i + 1
        b = b + 1
        # строка значений начинается со второй, следующей за строкой оглавления (row_
        row = row + 1

        while b < (count_element_from_dict + 1):
            print('\n----Анализ совпадений ГРЗ списка № ' + str(i) + ' со списком № ' + str(b) + '----')
            # заполняем ячейки оглавления
            cell_name = sheet.cell(row=row_number, column=col_number)
            cell_name.value = f"{i} и {b}"

            # Применяем set(множество) к списку ГРЗ для удаления повторяющихся ГРЗ.
            for el in set(dict_input[i]):
                if el in dict_input[b]:
                    print(str(el) + '\t из списка № ' + str(i) + '\t---> совпадение со списком № ' + str(b))
                    # заполняем ячейки повторяющимися ГРЗ
                    cell_value = sheet.cell(row=row, column=column)
                    cell_value.value = el
                    # Добавляем к строке единицу
                    row = row + 1
            # список № 2 становится списком № 3 и т.д.
            b = b + 1
            # столбец оглавления сдвигается вправо
            col_number = col_number + 1
            # столбец с совпадающими ГРЗ сдвигается вправо
            column = column + 1
            # строка снова становится на вторую позицию (следующую за строкой с оглавлением)
            row = 2
        # второй список превращается
        b = 1 + i
    wb.save('result.xlsx')
    return dict_input


def iteration_summ_list(dict_input_user):
    """Добавить описание"""
    # Создаем в открытом xlsx файле новый лист
    wb.create_sheet(title='Общее сравнение', index=1)
    sheet = wb['Общее сравнение']
    # Сравниваем мужду собой списоки a и b. Список b всегда больше списка a.
    a = 1
    b = 2
    # Создаем счетчик i, который после каждого завершения операции увеличивается на 1
    i = 0
    # номера ячеек для заголовка (строка + столбец)
    row_number = 1
    col_number = 1
    # номера ячеек для значений (строка + столбец)
    row = 1
    column = 1

    dict_input = dict_input_user
    while a < len(dict_input):
        print('------ СОВПАДЕНИЯ СПИСКА № ' + str(a) + ' -------')
        while i < len(dict_input) and b < (len(dict_input) + 1):
            # Если пересечения множества списков a и b имеет 1 и более элементов
            if len(set(dict_input[a]) & set(dict_input[b])) > 0:
                row = row + 1
                # заполняем ячейки оглавления
                cell_name = sheet.cell(row=row_number, column=col_number)
                cell_name.value = f"с {a} по {b}"

                print('Совпадения между списком № ' + str(a) + ' и списком № ' + str(b))
                print(set(dict_input[a]) & set(dict_input[b]))
                result = list(set(dict_input[a]) & set(dict_input[b]))

                for el in result:
                    # заполняем ячейки повторяющимися ГРЗ
                    cell_value = sheet.cell(row=row, column=column)
                    cell_value.value = el
                    # Добавляем к строке единицу
                    row = row + 1

                # столбец оглавления сдвигается вправо
                col_number = col_number + 1
                # столбец с совпадающими ГРЗ сдвигается вправо
                column = column + 1
                # строка снова становится на вторую позицию (следующую за строкой с оглавлением)
                row = 2
                b = b + 1
                while i < len(dict_input) and b < (len(dict_input) + 1):
                    if len(set(result) & set(dict_input[b])) > 0:
                        # столбец оглавления сдвигается вправо
                        col_number = col_number + 1
                        # столбец с совпадающими ГРЗ сдвигается вправо
                        column = column + 1
                        cell_name = sheet.cell(row=row_number, column=col_number)
                        cell_name.value = f"с {a} по {b}"

                        print('Совпадения предыдущих списков со списокм № ' + str(b))
                        print(set(result) & set(dict_input[b]))
                        result = list(set(result) & set(dict_input[b]))

                        for el in result:
                            # заполняем ячейки повторяющимися ГРЗ
                            cell_value = sheet.cell(row=row, column=column)
                            cell_value.value = el
                            # Добавляем к строке единицу
                            row = row + 1

                        b = b + 1
                    else:
                        print('Нет совпадений предыдущих списков со списокм № ' + str(b))
                        b = b + 1
                    i = i + 1
            else:
                print('Совпадения между списком № ' + str(a) + ' и списком № ' + str(b) + ' НЕТ')
                b = b + 1
        a = a + 1
        b = a + 1
        wb.save('result.xlsx')


dict_input, other = return_dict_all_grz()


first = iteration_one_to_many_lists(dict_input)
iteration_summ_list(dict_input)

print('Список номеров из других регионов, а также с нечитаемым регионом')
print(other)
